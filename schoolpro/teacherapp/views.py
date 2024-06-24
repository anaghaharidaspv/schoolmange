from datetime import datetime,timedelta
from cycler import V
from django.conf import settings
from django.core.mail import send_mail
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.utils.crypto import get_random_string
from django.views import View
import openpyxl
from adminapp.models import Teacher
from .forms import *
from django.contrib.auth import authenticate, login, logout
from django.http import HttpResponse
from django.views import View
from openpyxl import Workbook
from certificateapp.models import Reportcard ,Subject,Student
from django.contrib import messages


class TeacherLogin(View):
    def get(self, request):
        form = TeacherLoginForm()
        return render(request, 'teacherapp/teacherlogin.html', {'form': form})

    def post(self, request):
        msg = ''
        if request.method == 'POST':
            form = TeacherLoginForm(request.POST)
            if form.is_valid():
                username = form.cleaned_data['username']
                password = form.cleaned_data['password']
                print(username, password)
                teacher = Teacher.objects.get(username=username, password=password)
                # teacher=authenticate(username=username,password=password)
                print(teacher)
                if teacher is not None:
                    login(request, teacher)
                    return render(request, 'teacherapp/teacher_profile.html', {'teacher': teacher})
                else:
                    msg = 'error'
                    return render(request, 'teacherapp/teacherlogin.html', {'form': form, 'msg': msg})
        else:
            form = TeacherLoginForm()
        return render(request, 'teacherapp/teacherlogin.html', {'form': form})


class TeacherProfile(View):
    def get(self, request):
        return render(request, 'teacherapp/teacher_profile.html')


class TeacherLogout(View):
    def get(self, request):
        logout(request)
        return redirect('teacher_login')


class ChangePassword(View):
    def get(self, request, id):
        return render(request, 'teacherapp/update_password.html')

    def post(self, request, id):

        teacher_obj = Teacher.objects.get(id=id)
        if request.method == 'POST':
            old_password = request.POST.get('old_password')
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')
            if teacher_obj.password == old_password:
                if new_password == confirm_password:
                    teacher_obj.password = new_password
                    teacher_obj.save()
                    return HttpResponse('Password Reset Successfully')
                else:
                    msg = 'New Password and Confrim Password is not matching'
                    return render(request, 'teacherapp/update_password.html', {'msg': msg})
            else:
                msg = 'Current Password You Entered is Wrong'
                return render(request, 'teacherapp/update_password.html', {'msg': msg})


class LoginEmailOtp(View):
    def get(self, request):
        return render(request, 'teacherapp/otp.html')

    def post(self, request):
        if request.method == 'POST':
            email = request.POST.get('email')
            print(email)
            teacher = Teacher.objects.get(email=email)  # checking wheather teacher exist in given username and password
            if teacher is not None:
                request.session['teacher'] = teacher.id  # creating a sesseion for teacher to get teacher object in next View(OtpVerification(View))
                otp = get_random_string(length=6, allowed_chars='1234567890')
                expiration_time = datetime.now() + timedelta(minutes=1)
                expiration_time_str = expiration_time.strftime('%Y-%m-%d %H:%M:%S')
                subject = 'OTP for Login'
                message = f'Your OTP for login account : {otp}'
                from_email = settings.EMAIL_HOST_USER
                to_email_list = [email]
                send_mail(subject, message, from_email, to_email_list)
                request.session['otp'] = {'code': otp, 'expiration_time_string': expiration_time_str}
                return redirect('verifyotp')
            else:
                msg = 'user does not exist'
                return render(request, 'teacherapp/otp.html', {'msg': msg})
        else:
            return render(request, 'teacherapp/otp.html')




class OtpVerification(View):
    def get(self, request):
        return render(request, 'teacherapp/otp_verification.html')

    def post(self, request):
        if request.method == 'POST':
            otp_entered = request.POST.get('otp')
            otp_session = request.session.get('otp')
            if otp_session:
                
                
                saved_otp = otp_session.get('code')
                expiry_time = otp_session.get('expiration_time_string')
                expiration_time = datetime.strptime(expiry_time, '%Y-%m-%d %H:%M:%S')
                if datetime.now() <= expiration_time:
                    if saved_otp and otp_entered == saved_otp:
                        teacher_obj = request.session.get('teacher')
                        teacher = Teacher.objects.get(id=teacher_obj)
                        login(request, teacher)
                        return render(request, 'teacherapp/teacher_profile.html', {'teacher': teacher})
                    else:
                        msg = 'wrong otp'
                        return render(request, 'teacherapp/otp_verification.html', {'msg': msg})
                else:
                    msg1 = 'Otp expired'
                    return render(request, 'teacherapp/otp_verification.html', {'msg1': msg1})




class UploadFileView(View):
    def get(self, request):
        form = UploadFileForm()
        return render(request, 'files/upload_file.html', {'form': form})

    def post(self, request):
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('teacherapp:file_list')
        return render(request, 'files/upload_file.html', {'form': form})

class FileListView(View):
    def get(self, request):
        files = UploadFile.objects.all()
        return render(request, 'files/file_list.html', {'files': files})

class DeleteFileView(View):
    def get(self, request, pk):
        file = UploadFile.objects.get(pk=pk)
        file.delete()
        return redirect('teacherapp:file_list')
    
class ImportDataView(View):
    def get(self, request):
        form = Excelform()
        return render(request, 'sheet/excel_upload.html', {'form': form})

    def post(self, request):
        form = Excelform(request.POST, request.FILES)
        if form.is_valid():
            if 'file' not in request.FILES:
                messages.error(request, "No file uploaded!")
                return render(request, 'sheet/excel_upload.html', {'form': form})
            file = request.FILES['file']
            wb = openpyxl.load_workbook(file)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                Reportcard.objects.create(
                    stuname=Student.objects.get(first_name=row[0]),  # Assuming stuname is a ForeignKey to a Student model
                    sub1=Subject.objects.get(name=row[1]),           # Assuming sub1 is a ForeignKey to a Subject model
                    mark1=row[2],
                    sub2=Subject.objects.get(name=row[3]),
                    mark2=row[4],
                    sub3=Subject.objects.get(name=row[5]),
                    mark3=row[6],
                    sub4=Subject.objects.get(name=row[7]),
                    mark4=row[8],
                    sub5=Subject.objects.get(name=row[9]),
                    mark5=row[10]
                )
            messages.success(request, "Data imported successfully!")
            return HttpResponse("imported successfully")
        return render(request, 'sheet/excel_upload.html', {'form': form})

class ExportToExcelView(View):
    def get(self, request, *args, **kwargs):
        # Create a workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Excel Upload Data"
        # Define the headers
        headers = ['stuname', 'sub1', 'mark1','sub2','mark2','sub3','mark3','sub4','mark4','sub5','mark5']
        ws.append(headers)
        # Fetch the data from the database
        excel_uploads = Reportcard.objects.all()
        # Append the data to the worksheet
        for upload in excel_uploads:
            ws.append([upload.stuname.first_name, upload.sub1.name,upload.mark1,upload.sub2.name,upload.mark2,upload.sub3.name,upload.mark3,upload.sub4.name,upload.mark4,upload.sub5.name,upload.mark5])
        # Create an HTTP response with the appropriate headers for Excel file download
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=excelupload_data.xlsx'
        # Save the workbook to the response
        wb.save(response)
        return response


class AchievementListView(View):
    def get(self, request):
        achievements = Achievement.objects.all()
        return render(request, 'achievements/achievements_list.html', {'achievements': achievements})

class AchievementCreateView(View):
    def get(self, request):
        form = AchievementForm()
        return render(request, 'achievements/achievements_form.html', {'form': form})

    def post(self, request):
        form = AchievementForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('achievement-list')
        return render(request, 'achievements/achievements_form.html', {'form': form})

class  Editachievement(View):
    def get(self,request,pk):
        ach_obj=Achievement.objects.get(pk=pk)
        form = AchievementForm(instance=ach_obj)
        return render(request, 'achievements/achievements_form.html',{'form':form})

    def post(self, request,pk):
        ach_obj=Achievement.objects.get(pk=pk)
        form = AchievementForm(request.POST,instance=ach_obj)
        if form.is_valid():
            form.save()
            return redirect('achievement-list')
        return render(request, 'achievements/achievements_form.html', {'form': form})

class  Deleteachievement(View):
    def get(self,request,pk):
        ach_obj=Achievement.objects.get(pk=pk)
        ach_obj.delete()
        return redirect('achievement-list')
